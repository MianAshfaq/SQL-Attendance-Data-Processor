import pandas as pd
import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import messagebox, simpledialog
from tkcalendar import DateEntry

# Function to connect to the SQL database using a direct connection string
def connect_to_database(server, database, username, password):
    try:
        conn_str = f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};TrustServerCertificate=yes;'
        conn = pyodbc.connect(conn_str)
        print("Connected to the SQL database.")
        return conn
    except Exception as e:
        print(f"Connection failed: {e}")
        return None

# Function to retrieve the list of employees from the SQL database
def retrieve_employee_list(conn):
    try:
        query = "SELECT Badgenumber, Name FROM dbo.USERINFO"
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        employees = {row[0]: row[1] for row in rows}
        print("Employee list retrieved.")
        return employees
    except Exception as e:
        print(f"Employee list retrieval failed: {e}")
        return {}

# Function to retrieve attendance data from the SQL database
def retrieve_attendance_data(conn, start_date, end_date, selected_ids):
    try:
        ids_filter = ''
        if selected_ids:
            ids_filter = f"AND USERINFO.Badgenumber IN ({', '.join([f'\'{id.strip()}\'' for id in selected_ids.split(',')])})"
        
        query = f"""
        SELECT CHECKINOUT.USERID, CHECKINOUT.CHECKTIME, CHECKINOUT.CHECKTYPE, USERINFO.Badgenumber, USERINFO.Name AS EmployeeName
        FROM dbo.CHECKINOUT
        LEFT JOIN dbo.USERINFO ON CHECKINOUT.USERID = USERINFO.USERID
        WHERE CHECKINOUT.CHECKTIME >= '{start_date}' AND CHECKINOUT.CHECKTIME < '{end_date}'
        {ids_filter}
        """
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        df = pd.DataFrame.from_records(rows, columns=columns)
        print("Attendance data retrieved.")
        return df
    except Exception as e:
        print(f"Data retrieval failed: {e}")
        return None

# Function to process and save attendance data into Excel sheets
def process_and_save_data(df, file_path):
    try:
        if df.empty:
            print("No attendance data to save.")
            return
        
        df['CHECKTIME'] = pd.to_datetime(df['CHECKTIME'])
        df['Date'] = df['CHECKTIME'].dt.strftime('%m/%d/%Y')  # Format date as MM/DD/YYYY
        df['Time'] = df['CHECKTIME'].dt.time

        # Create a sequential count of events for each user per day
        df['Event'] = df.groupby(['USERID', 'Date']).cumcount() + 1

        # Map event numbers to desired IN/OUT order
        def map_event(event):
            if event == 1:
                return 'IN_1'
            elif event == 2:
                return 'OUT_1'
            elif event == 3:
                return 'IN_2'
            elif event == 4:
                return 'OUT_2'
            elif event == 5:
                return 'IN_3'
            elif event == 6:
                return 'OUT_3'
            else:
                return f'EXTRA_{event}'

        df['MappedEvent'] = df['Event'].apply(map_event)

        # Pivot the table to get separate columns for each mapped event
        df_pivot = df.pivot_table(index=['Date', 'Badgenumber', 'EmployeeName'], columns=['MappedEvent'], values='Time', aggfunc='first')
        
        # Flatten the column names
        df_pivot.columns = [col for col in df_pivot.columns]

        # Define the desired column order
        desired_columns = ['IN_1', 'OUT_1', 'IN_2', 'OUT_2', 'IN_3', 'OUT_3']

        # Add missing columns if necessary to maintain the order
        for col in desired_columns:
            if col not in df_pivot.columns:
                df_pivot[col] = None

        df_pivot = df_pivot[desired_columns]

        df_pivot.reset_index(inplace=True)
        df_pivot.rename(columns={'Badgenumber': 'ID Number', 'Date': 'Date', 'EmployeeName': 'Name'}, inplace=True)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for employee_id, employee_data in df_pivot.groupby('ID Number'):
                employee_name = employee_data['Name'].iloc[0]
                sheet_name = (f'{employee_name}_{employee_id}')[:20]
                employee_data.to_excel(writer, sheet_name=sheet_name, index=False)

        # Load the workbook and apply formatting
        wb = load_workbook(file_path)
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for sheet in wb.worksheets:
            # Apply header styles
            header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            header_font = Font(bold=True, color="000000", name="SimSun", size=11)  # Black text, bold, SimSun font, size 11
            header_alignment = Alignment(horizontal="center", vertical="center")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border  # Add border to header cells

            # Apply alternating row styles, borders, and font styles
            even_row_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if cell.row % 2 == 0:
                        cell.fill = even_row_fill
                    cell.border = thin_border  # Add border to all cells
                    cell.font = Font(name="SimSun", size=11)  # Set font to SimSun, size 11

            # Auto-adjust column widths
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            # Rename columns to "IN" and "OUT"
            for col_num in range(4, sheet.max_column + 1, 2):
                sheet.cell(row=1, column=col_num).value = "IN"
                sheet.cell(row=1, column=col_num + 1).value = "OUT"
        
        wb.save(file_path)
        print("Data saved to Excel sheets with formatting.")
    except Exception as e:
        print(f"Data processing and saving failed: {e}")

# Function to get the date range and selected IDs from the user using a popup with a calendar and listbox
def get_date_range_and_ids(employee_dict):
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    date_range = {}
    selected_ids = []

    def get_data():
        nonlocal selected_ids
        date_range['start_date'] = start_cal.get_date().strftime('%Y-%m-%d')
        date_range['end_date'] = end_cal.get_date().strftime('%Y-%m-%d')
        selected = listbox.curselection()
        selected_ids = [listbox.get(i).split(':')[0].strip() for i in selected]
        root.destroy()

    def select_all_employees():
        listbox.select_set(0, tk.END)

    def enter_manual_ids():
        nonlocal selected_ids
        manual_ids = simpledialog.askstring("Manual IDs", "Enter employee IDs separated by comma:")
        if manual_ids:
            selected_ids = [id.strip() for id in manual_ids.split(',')]
            # Update date range with default values
            date_range['start_date'] = start_cal.get_date().strftime('%Y-%m-%d')
            date_range['end_date'] = end_cal.get_date().strftime('%Y-%m-%d')
        root.destroy()

    popup = tk.Toplevel()
    popup.title("Select Date Range and Employee IDs")
    popup.configure(background='skyblue')  # Set background color to sky blue

    tk.Label(popup, text="Start Date:", bg='skyblue').grid(row=0, column=0, padx=10, pady=10)
    start_cal = DateEntry(popup, width=12, background='darkblue', foreground='white', borderwidth=2)
    start_cal.grid(row=0, column=1, padx=10, pady=10)

    tk.Label(popup, text="End Date:", bg='skyblue').grid(row=1, column=0, padx=10, pady=10)
    end_cal = DateEntry(popup, width=12, background='darkblue', foreground='white', borderwidth=2)
    end_cal.grid(row=1, column=1, padx=10, pady=10)

    tk.Label(popup, text="Employee IDs:", bg='skyblue').grid(row=2, column=0, padx=10, pady=10)
    listbox = tk.Listbox(popup, selectmode=tk.MULTIPLE, width=30, height=10)
    for key, value in employee_dict.items():
        listbox.insert(tk.END, f"{key}: {value}")
    listbox.grid(row=2, column=1, padx=10, pady=10)

    tk.Button(popup, text="Select All", command=select_all_employees).grid(row=3, column=0, columnspan=2, pady=5)
    tk.Button(popup, text="Enter Manual IDs", command=enter_manual_ids).grid(row=4, column=0, columnspan=2, pady=5)
    tk.Button(popup, text="Submit", command=get_data).grid(row=5, column=0, columnspan=2, pady=10)

    # Set minimum size for the popup
    popup.geometry("350x400")

    root.mainloop()
    return date_range['start_date'], date_range['end_date'], ','.join(selected_ids)

# Main script here add server IP,DB name,Userbame, & Password
try:
    server = '192.168.1.XXXX'
    database = 'XXXXXXX'
    username = 'XXX'
    password = 'XXX'

    conn = connect_to_database(server, database, username, password)
    if conn:
        employees = retrieve_employee_list(conn)
        start_date, end_date, selected_ids = get_date_range_and_ids(employees)
        print(f"Selected Date Range: {start_date} to {end_date}")
        print(f"Selected Employee IDs: {selected_ids}")
        
        df = retrieve_attendance_data(conn, start_date, end_date, selected_ids)
        if df is not None and not df.empty:
            file_path = 'attendance_data_v7.xlsx'
            process_and_save_data(df, file_path)
        else:
            print("No data found for the given criteria.")
    else:
        print("Failed to connect to the database.")
except Exception as e:
    print(f"An error occurred: {e}")
    messagebox.showerror("Error", f"An error occurred: {e}")

input("Press Enter to exit...")
